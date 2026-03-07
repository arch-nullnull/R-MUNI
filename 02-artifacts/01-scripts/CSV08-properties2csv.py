#!/usr/bin/env python3
# CSV08-properties2csv.py
#
# Purpose:
#   Append Properties from XLSX exports into existing master properties.csv
#
# Scope:
#   - Only SOURCE=XLSX entries from run-scope.txt are processed
#   - Only the XLSX sheet named "Properties" is evaluated
#   - Mapping is explicit via propmapping.txt
#
# Behavior:
#   - Root resolved via CSV00-root.resolved.txt
#   - Append-only (no overwrite, no delete)
#   - Duplicate keys allowed; newest entry wins by position
#   - Owner ID must exist in elements.csv or relations.csv
#
# Mapping:
#   <root>/02-artifacts/03-XLSX/01-mapping/propmapping.txt
#
# Logging:
#   <root>/03-stages/99-logs/CSV08-properties2csv.log

import os
import sys
import csv
from datetime import datetime
from openpyxl import load_workbook


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def die(msg):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def log(msg, log_path):
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[CSV08] {now()} | {msg}\n")


def read_csv_header(path):
    with open(path, newline="", encoding="utf-8") as f:
        return next(csv.reader(f))


# ------------------------------------------------------------
# Root resolution (CSV00)
# ------------------------------------------------------------

def resolve_root():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_file = os.path.join(
        script_dir, "..", "..", "03-stages", "99-logs", "CSV00-root.resolved.txt"
    )

    if not os.path.isfile(root_file):
        die(f"Missing root resolution file: {root_file}")

    with open(root_file, encoding="utf-8") as f:
        root = f.readline().strip()

    if not root or not os.path.isdir(root):
        die(f"Invalid root resolved: {root}")

    return root


# ------------------------------------------------------------
# Run-scope handling (identical semantics to CSV07)
# ------------------------------------------------------------

def read_run_scope_xlsx(path):
    models = []
    current_source = None

    if not os.path.isfile(path):
        return models

    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if line.startswith("SOURCE="):
                current_source = line.split("=", 1)[1]

            elif line.startswith("MODEL=") and current_source == "XLSX":
                model = line.split("=", 1)[1]
                if not model.startswith("~$"):
                    models.append(model)

    return models


# ------------------------------------------------------------
# Mapping loader
# ------------------------------------------------------------

def load_propmapping(path):
    """
    propmapping.txt format:
    "XLSX_Header","Sheet","CSV_Header","properties.csv"
    """
    mapping = {}

    with open(path, encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or row[0].startswith("#"):
                continue

            xlsx_col, sheet, csv_col, target = row
            if target != "properties.csv":
                continue

            mapping[xlsx_col] = csv_col

    for required in ("Owner Id", "Key", "Value"):
        if required not in mapping:
            die(f"Missing mapping for '{required}' in propmapping.txt")

    return mapping


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def main():

    root = resolve_root()

    log_dir = os.path.join(root, "03-stages", "99-logs")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "CSV08-properties2csv.log")

    log("=== CSV08 RUN START ===", log_path)

    # Paths
    run_scope_path = os.path.join(root, "03-stages", "run-scope.txt")
    xlsx_root = os.path.join(
        root, "02-artifacts", "03-XLSX", "03-child", "00-archimatechild"
    )
    mapping_path = os.path.join(
        root, "02-artifacts", "03-XLSX", "01-mapping", "propmapping.txt"
    )
    master_csv_dir = os.path.join(root, "02-artifacts", "02-csv", "00-master")

    properties_csv = os.path.join(master_csv_dir, "properties.csv")
    elements_csv = os.path.join(master_csv_dir, "elements.csv")
    relations_csv = os.path.join(master_csv_dir, "relations.csv")

    # Validate master CSVs
    for p in (properties_csv, elements_csv, relations_csv):
        if not os.path.isfile(p):
            die(f"Missing required master CSV: {p}")

    # Load valid owner IDs
    valid_ids = set()
    for path in (elements_csv, relations_csv):
        header = read_csv_header(path)
        id_idx = header.index("ID")
        with open(path, encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                if row and row[id_idx]:
                    valid_ids.add(row[id_idx])

    # Load mapping
    mapping = load_propmapping(mapping_path)
    log("Property mapping loaded", log_path)

    # Resolve XLSX scope
    xlsx_files = read_run_scope_xlsx(run_scope_path)
    if not xlsx_files:
        log("No SOURCE=XLSX entries found – nothing to do", log_path)
        return

    log(f"XLSX files resolved: {xlsx_files}", log_path)

    appended = 0

    # Open properties.csv for append
    with open(properties_csv, "a", newline="", encoding="utf-8") as out:
        writer = csv.writer(out, quoting=csv.QUOTE_ALL)

        for fname in xlsx_files:
            xlsx_path = os.path.join(xlsx_root, fname)
            if not os.path.isfile(xlsx_path):
                die(f"XLSX not found: {fname}")

            log(f"Processing XLSX: {fname}", log_path)
            wb = load_workbook(xlsx_path, data_only=True)

            if "Properties" not in wb.sheetnames:
                log(f"{fname} | sheet missing: Properties", log_path)
                continue

            sheet = wb["Properties"]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header = list(rows[0])
            data_rows = rows[1:]

            for row in data_rows:
                if not any(row):
                    continue

                source = dict(zip(header, row))
                owner_id = str(source.get("Owner Id", "")).strip()
                key = str(source.get("Key", "")).strip()
                value = str(source.get("Value", "")).strip()

                if not owner_id or not key:
                    continue

                if owner_id not in valid_ids:
                    die(f"Owner ID not found in master CSVs: {owner_id}")

                writer.writerow([owner_id, key, value])
                appended += 1

    log(f"Properties appended: {appended}", log_path)
    log("CSV08 completed successfully", log_path)
    print("[CSV08] OK | properties2csv completed")


if __name__ == "__main__":
    main()
