#!/usr/bin/env python3
# CSV07-xlsx_2_csv.py
#
# Purpose:
#   Import XLSX-based Archi exports into existing master CSVs
#   using an explicit csvmapping.txt.
#
# Scope:
#   - Only mappings defined in csvmapping.txt are processed
#   - No implicit sheet or column handling
#   - No Properties / Specializations logic
#
# Behavior:
#   - Scope-aware via run-scope.txt
#   - Fallback XLSX discovery if no XLSX scope is defined
#   - Column order enforced from target CSV
#   - Append-only
#
# Mapping:
#   <root>/01-artifacts/03-XLSX/01-mapping/csvmapping.txt
#
# Logging:
#   <root>/02-stages/99-logs/CSV07-xlsx_2_csv.log

import os
import csv
from datetime import datetime
from openpyxl import load_workbook


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def log(msg, log_path):
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[CSV07] {now()} | {msg}\n")


def read_csv_header(path):
    with open(path, newline="", encoding="utf-8") as f:
        return next(csv.reader(f))


# ------------------------------------------------------------
# Scope handling
# ------------------------------------------------------------

def read_run_scope(path):
    if not os.path.isfile(path):
        return []

    models = []
    current_source = None

    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue

            if line.startswith("SOURCE="):
                current_source = line.split("=", 1)[1]

            elif line.startswith("MODEL=") and current_source == "XLSX":
                model = line.split("=", 1)[1]
                if not model.startswith("~$"):
                    models.append(model)

    return models


def discover_xlsx(folder):
    return sorted(
        f for f in os.listdir(folder)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    )


# ------------------------------------------------------------
# Mapping loader
# ------------------------------------------------------------

def load_mapping(path):
    """
    csvmapping.txt format:
    "XLSX_Header","Sheet","CSV_Header","target.csv"
    """
    mapping = {}

    with open(path, encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or row[0].startswith("#"):
                continue

            xlsx_header, sheet, csv_header, target_csv = row

            mapping.setdefault(sheet, {}).setdefault(target_csv, []).append(
                (xlsx_header, csv_header)
            )

    return mapping


# ------------------------------------------------------------
# Main
# ------------------------------------------------------------

def main():

    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_file = os.path.join(
        script_dir, "..", "..", "02-stages", "99-logs", "CSV00-root.resolved.txt"
    )

    with open(root_file, encoding="utf-8") as f:
        root = f.readline().strip()

    log_dir = os.path.join(root, "02-stages", "99-logs")
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "CSV07-xlsx_2_csv.log")

    log("=== CSV07 RUN START ===", log_path)

    # Paths
    run_scope_path = os.path.join(root, "02-stages", "run-scope.txt")
    xlsx_root = os.path.join(
        root, "01-artifacts", "03-XLSX", "03-child", "00-archimatechild"
    )
    mapping_path = os.path.join(
        root, "01-artifacts", "03-XLSX", "01-mapping", "csvmapping.txt"
    )
    master_csv_dir = os.path.join(root, "01-artifacts", "02-csv", "00-master")

    # Load mapping
    mapping = load_mapping(mapping_path)
    log(f"Mapping loaded for sheets: {list(mapping.keys())}", log_path)

    # Resolve XLSX files
    scoped = read_run_scope(run_scope_path)
    if scoped:
        log("run-scope.txt found – XLSX scope entries detected", log_path)
        xlsx_files = scoped
    else:
        log("No XLSX scope entries found – fallback discovery active", log_path)
        xlsx_files = discover_xlsx(xlsx_root)

    log(f"XLSX files resolved: {xlsx_files}", log_path)

    # Process XLSX files
    for fname in xlsx_files:
        xlsx_path = os.path.join(xlsx_root, fname)
        if not os.path.isfile(xlsx_path):
            log(f"SKIPPED missing XLSX: {fname}", log_path)
            continue

        log(f"Processing XLSX: {fname}", log_path)
        wb = load_workbook(xlsx_path, data_only=True)

        for sheet_name, targets in mapping.items():
            if sheet_name not in wb.sheetnames:
                log(f"{fname} | sheet missing: {sheet_name}", log_path)
                continue

            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            source_header = list(rows[0])
            data_rows = rows[1:]

            for target_csv, column_map in targets.items():
                target_path = os.path.join(master_csv_dir, target_csv)
                target_header = read_csv_header(target_path)

                records = []
                for row in data_rows:
                    if not any(row):
                        continue

                    source = dict(zip(source_header, row))
                    record = {key: "" for key in target_header}

                    for xlsx_col, csv_col in column_map:
                        if csv_col in record:
                            record[csv_col] = source.get(xlsx_col, "")

                    records.append(record)

                if records:
                    with open(target_path, "a", newline="", encoding="utf-8") as f:
                        writer = csv.DictWriter(
                            f,
                            fieldnames=target_header,
                            quoting=csv.QUOTE_ALL
                        )
                        for r in records:
                            writer.writerow(r)

                log(
                    f"{fname} | {sheet_name} -> {target_csv} | {len(records)} rows appended",
                    log_path
                )

    log("CSV07 completed successfully", log_path)
    print("[CSV07] OK | xlsx_2_csv completed")


if __name__ == "__main__":
    main()
